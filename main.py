from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import *
from time import sleep
import openpyxl
import smtplib
import os
from email.message import EmailMessage
import re


class Scrappy:

    def iniciar(self):
        self.verificar_email_usuario()
        self.configurar_chromedriver()
        self.raspagem_de_dados()
        self.criar_planilha()
        self.enviar_email_cliente()

    def verificar_email_usuario(self):
        while True:
            self.email = input(
                'Digite o email para receber o relatorio de valores dos celulares!\n')
            self.email.lower()

            valido = re.search(
                r'[a-zA-Z0-9_-]+@[a-zA-Z0-9]+\.[a-zA-Z]{1,3}$', self.email)
            if valido:
                print(f'email {self.email} válido')
                break

            
    
    def configurar_chromedriver(self):
        chrome_options = Options()
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        chrome_options.add_argument('--lang=pr-BR')
        chrome_options.add_argument('--disable-notifications')
        self.driver = webdriver.Chrome(executable_path='chromedriver.exe')
        self.driver.set_window_size(800, 700)
        self.driver.get('https://www.youtube.com/?app=desktop&gl=BR')


    def raspagem_de_dados(self):
        self.link = 'https://telefonesimportados.netlify.app/'
        print(self.driver.title)
        self.lista_nome_celulares = []
        self.lista_preco_celulares = []
        self.driver.get(self.link)
        sleep(2)
        
        while True:
            item = 1
            lista_nomes = self.driver.find_elements_by_xpath(
                f'/html/body/div[5]/div[2]/div[1]/div[{item}]/div/h2/a')
            lista_precos = self.driver.find_elements_by_xpath(
                f'//div[{item}]/div[@class="single-shop-product" and 1]/div[@class="product-carousel-price" and 2]/ins[1]')
            if lista_nomes and lista_precos:
                print("Achei um novo celular!")
                self.lista_nome_celulares.append(lista_nomes[0].text)
                sleep(1)
                self.lista_preco_celulares.append(lista_precos[0].text)
                print("Adicionei um novo celular na lista")
                item += 1
                sleep(1)
            else:
                botao_proximo = self.driver.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/nav/ul/li[7]/a')
                if botao_proximo:
                    botao_proximo.click()
                    print(f'\u001b[32m{"Navegando para proxima pagina"}\u001b[0m')
                    
                else:
                    print(f'\u001b[33m{"Não há mais paginas!"}\u001b[0m')
                    print(f'\u001b[32m{"Escaneamento Concluido"}\u001b[0m')
                    self.driver.quit()


    def criar_planilha(self):
        index = 2
        planilha = openpyxl.Workbook()
        celulares = planilha['Sheet']
        celulares.title = 'Celulares'
        celulares['A1'] = 'Nome'
        celulares['B1'] = 'Preço'
        for nome, preco in zip(self.lista_nome_celulares, self.lista_preco_celulares):
            celulares.cell(column=1, row=index, value=nome)
            celulares.cell(column=2, row=index, value=preco)
            index += 1
        planilha.save("planilha_de_preços.xlsx")

        print(f'\u001b[32m{"Planilha criada com sucesso"}\u001b[0m')

    def enviar_email_cliente(self, endereco, senha):
        msg = EmailMessage()
        msg['Subject'] = 'planilha de Preços de Telefones Importados'
        msg['From'] = endereco
        msg['To'] = self.email
        msg.set_content('Olá a sua planilha chegou')
        arquivos = ["planilha_de_preços.xlsx"]
        for arquivo in arquivos:
            with open(arquivo, 'rb') as arq:
                dados = arq.read()
                nome_arquivo = arq.name
            msg.add_attachment(dados, maintype='application',
                               subtype='octet-stream', filename=nome_arquivo)
        server = smtplib.SMTP('imap.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(endereco, senha, initial_response_ok=True)
        server.send_message(msg)
        print(f'\u001b[32m{"Enviando email para para destinatario"}\u001b[0m')

        server.quit()


start = Scrappy()
start.iniciar()
